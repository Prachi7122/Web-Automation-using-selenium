import openpyxl
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)
#
# def readData(file,sheetName,rownum,columnno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.get_sheet_by_name(sheetName)
#     return sheet.cell(row=rownum,column=columnno).value

# import openpyxl

#
# def readData(file, sheetName, rownum, columnName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#
#     # Convert the column name to column number
#     columnno = openpyxl.utils.column_index_from_string(columnName)
#
#     # Your implementation to read data from the specified row and column
#     data = sheet.cell(row=rownum, column=columnno).value
#
#     workbook.close()
#
#     print("Column Name:", columnName)  # You can remove this line if not needed
#
#     return data

#
# def readData(file, sheetName, rownum, columnName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#
#     # Convert the column name to column number
#     columnno = openpyxl.utils.column_index_from_string(columnName)
#
#     # Your implementation to read data from the specified row and column
#     data = sheet.cell(row=rownum, column=columnno).value
#
#     workbook.close()
#
#     print("Column Name:", columnName)
#     columnno = openpyxl.utils.column_index_from_string(columnName)
#
#     return data


def get_column_index(sheet, column_name):
    for col_idx, col in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1)):
        if col[0].value == column_name:
            return col_idx + 1  # Adding 1 to convert from 0-based index to 1-based index
    return None  # Return None if the column name is not found

def readData(file, sheet_name, row_num, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    column_index = get_column_index(sheet, column_name)

    if column_index is not None:
        return sheet.cell(row=row_num, column=column_index).value
    else:
        return None

def writeData(file, sheetName, rownum, columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_Sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)